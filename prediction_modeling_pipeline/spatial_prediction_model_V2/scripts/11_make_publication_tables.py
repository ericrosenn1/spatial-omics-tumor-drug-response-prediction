"""
Script: 11_make_publication_tables.py

Purpose:
    Create publication-ready tables and supporting-file packages from the Step
    10 package.

Pipeline role:
    This step converts the integrated interpretation package into short
    publication tables, an Excel workbook, copied supporting source files,
    manifests, reports, and a ZIP archive.

Scientific role:
    Publication tables make the V2 results reviewable by separating model
    comparison, validated treatments, recurrent spatial features, recurrent
    biology themes, figure metadata, recommendations, and provenance into
    manuscript-friendly artifacts.

Documentation polish marker:
    SPATIAL_PREDICTION_MODEL_V2_STEP11_DOC_POLISH_V1

Important:
    This documentation pass is intentionally non-behavioral. Comments,
    section headers, and docstrings may be added, but executable logic,
    imports, constants, thresholds, hyperparameters, validation rules,
    output filenames, and return codes must remain unchanged.
"""


# =============================================================================
# Imports and local package setup
# =============================================================================

from __future__ import annotations

import argparse
import hashlib
import json
import os
import shutil
import sys
import zipfile
from pathlib import Path

import pandas as pd


# =============================================================================
# Helper functions
# =============================================================================

def ensure_dir(path: Path | str) -> Path:
    """Create a directory if needed and return it as a Path."""

    path = Path(path)
    path.mkdir(parents=True, exist_ok=True)
    return path


def read_table(path: Path | str) -> pd.DataFrame:
    """Read a CSV or TSV table, returning an empty DataFrame when unavailable."""

    path = Path(path)

    if not path.exists() or path.stat().st_size == 0:
        return pd.DataFrame()

    try:
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path)
        return pd.read_csv(path, sep="\t")
    except Exception:
        return pd.DataFrame()


def write_table(df: pd.DataFrame, path: Path | str) -> Path:
    """Write a CSV or TSV table with parent-directory creation."""

    path = Path(path)
    ensure_dir(path.parent)

    if df is None:
        df = pd.DataFrame()

    if path.suffix.lower() == ".csv":
        df.to_csv(path, index=False)
    else:
        df.to_csv(path, sep="\t", index=False)

    return path


def write_json(obj: dict, path: Path | str) -> Path:
    """Write a JSON artifact with stable formatting."""

    path = Path(path)
    ensure_dir(path.parent)
    path.write_text(json.dumps(obj, indent=2, sort_keys=True), encoding="utf-8")
    return path


def write_text_report(path: Path | str, body: str) -> Path:
    """Write a text report with a filepath header."""

    path = Path(path)
    ensure_dir(path.parent)
    path.write_text(f"FILEPATH: {path}\n\n{body}", encoding="utf-8")
    return path


def terminal_block(title: str, lines: list[str]) -> str:
    """Format terminal output as a readable status block."""

    bar = "=" * 90
    return "\n".join([bar, title, bar] + lines)


def safe_filename(text: str, max_len: int = 120) -> str:
    """Convert text into a filesystem-safe filename."""

    text = str(text)
    out = []
    for ch in text:
        if ch.isalnum() or ch in "._-":
            out.append(ch)
        else:
            out.append("_")
    name = "".join(out).strip("_")
    while "__" in name:
        name = name.replace("__", "_")
    if len(name) > max_len:
        digest = hashlib.md5(name.encode("utf-8")).hexdigest()[:10]
        name = name[: max_len - 11] + "_" + digest
    return name or "unnamed"


def load_json(path: Path | str) -> dict:
    """Load a JSON file, returning an empty dictionary when unavailable."""

    path = Path(path)

    if not path.exists() or path.stat().st_size == 0:
        return {}

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def file_manifest(root: Path) -> pd.DataFrame:
    """Build a file manifest for all files under a root directory."""

    rows = []

    if not root.exists():
        return pd.DataFrame(columns=["relative_path", "path", "size_bytes"])

    for path in sorted(root.rglob("*")):
        if path.is_file():
            rows.append({
                "relative_path": str(path.relative_to(root)),
                "path": str(path),
                "size_bytes": int(path.stat().st_size),
            })

    return pd.DataFrame(rows)


def short_publication_table(df: pd.DataFrame, preferred_cols: list[str], max_rows: int | None = None) -> pd.DataFrame:
    """Select publication-facing columns and optionally truncate rows."""

    if df.empty:
        return pd.DataFrame()

    cols = [c for c in preferred_cols if c in df.columns]
    if not cols:
        cols = list(df.columns[: min(12, len(df.columns))])

    out = df[cols].copy()

    if max_rows is not None:
        out = out.head(max_rows)

    return out


def create_excel_workbook(tables: dict[str, pd.DataFrame], xlsx_path: Path) -> tuple[bool, str]:
    """Create and style a multi-sheet Excel workbook for publication tables."""

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except Exception as exc:
        return False, f"openpyxl import failed: {exc}"

    try:
        ensure_dir(xlsx_path.parent)

        sheet_rows = []
        for name, df in tables.items():
            sheet_rows.append({
                "sheet_name": name,
                "rows": int(len(df)),
                "columns": int(len(df.columns)) if not df.empty else 0,
                "description": table_description(name),
            })

        sheet_summary = pd.DataFrame(sheet_rows)
        all_tables = {"Sheet_Summary": sheet_summary} | tables

        with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
            for sheet_name, df in all_tables.items():
                safe_sheet = sheet_name[:31]
                if df.empty:
                    df = pd.DataFrame({"note": ["No rows available for this table."]})
                df.to_excel(writer, sheet_name=safe_sheet, index=False)

        wb = load_workbook(xlsx_path)

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.border = border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    cell.border = border

            for col_cells in ws.columns:
                values = [str(cell.value) for cell in col_cells if cell.value is not None]
                width = min(max([len(v) for v in values] + [10]) + 2, 45)
                ws.column_dimensions[col_cells[0].column_letter].width = width

            for row in range(1, min(ws.max_row, 200) + 1):
                ws.row_dimensions[row].height = 18

            ws.sheet_view.showGridLines = False

        wb.save(xlsx_path)
        return True, "excel_created"

    except Exception as exc:
        return False, str(exc)


def table_description(name: str) -> str:
    """Return a short reviewer-facing description for a workbook sheet."""

    descriptions = {
        "Model_Comparison": "Model family comparison with purpose, inputs, outputs, metrics, and validation status.",
        "Validated_Treatments": "Treatment-specific spatial biology findings after label shuffle validation.",
        "Recurrent_Features": "Spatial features recurring across model branches.",
        "Biology_Themes": "Recurrent biology themes across model branches.",
        "Figure_Manifest": "Presentation figure file list and captions.",
        "Integration_Recs": "Pipeline integration recommendations.",
        "Provenance": "Source and output provenance.",
        "Supporting_Files": "Short-path copied figures, source tables, reports, PDFs, and documentation artifacts included in the package.",
    }
    return descriptions.get(name, "Publication support table.")


def copy_supporting_files(step10_root: Path, support_root: Path) -> pd.DataFrame:
    """Copy figures, tables, reports, and source artifacts into short-path support folders."""

    figures_dir = ensure_dir(support_root / "copied_figures")
    source_dir = ensure_dir(support_root / "copied_source_tables_and_reports")

    rows = []

    for path in sorted(step10_root.rglob("*")):
        if not path.is_file():
            continue

        rel = path.relative_to(step10_root)
        suffix = path.suffix.lower()

        if suffix in [".png", ".jpg", ".jpeg"]:
            dst = figures_dir / safe_filename(str(rel), max_len=140)
            shutil.copy2(path, dst)
            rows.append({
                "source_path": str(path),
                "copied_path": str(dst),
                "file_type": "figure",
                "size_bytes": int(dst.stat().st_size),
            })

        elif suffix in [".tsv", ".csv", ".json", ".txt", ".pdf", ".xlsx"]:
            dst = source_dir / safe_filename(str(rel), max_len=140)

            if suffix == ".txt":
                try:
                    body = path.read_text(encoding="utf-8")
                except Exception:
                    body = path.read_text(errors="replace")
                dst.write_text(f"FILEPATH: {dst}\nSOURCE_FILEPATH: {path}\n\n{body}", encoding="utf-8")
            else:
                shutil.copy2(path, dst)

            rows.append({
                "source_path": str(path),
                "copied_path": str(dst),
                "file_type": "source_table_or_report",
                "size_bytes": int(dst.stat().st_size),
            })

    manifest = pd.DataFrame(rows)
    write_table(manifest, support_root / "supporting_source_file_manifest.tsv")

    readme = []
    readme.append("SUPPORTING SOURCE FILES")
    readme.append("")
    readme.append("This folder contains short-path copies of the Step 10 integrated interpretation source files and figures.")
    readme.append("Copied .txt files are wrapped with FILEPATH and SOURCE_FILEPATH headers.")
    readme.append("Use supporting_source_file_manifest.tsv to map copied files back to their original source paths.")
    write_text_report(support_root / "README_supporting_source_files.txt", "\n".join(readme))

    return manifest


def make_zip(output_root: Path, zip_path: Path) -> Path:
    """Create a ZIP archive for publication tables and supporting files."""

    ensure_dir(zip_path.parent)

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(output_root.rglob("*")):
            if not path.is_file():
                continue
            if path.resolve() == zip_path.resolve():
                continue
            zf.write(path, arcname=str(path.relative_to(output_root)))

    return zip_path


def maybe_open(path: Path):
    """Open a path on Windows when requested by the user."""

    try:
        if sys.platform.startswith("win"):
            os.startfile(str(path))
    except Exception:
        pass


# =============================================================================
# Main workflow
# =============================================================================

def main() -> int:
    """Run this spatial_prediction_model_V2 step and write tables, reports, provenance, and summaries."""

    parser = argparse.ArgumentParser()
    parser.add_argument("--step10-root", required=True)
    parser.add_argument("--output-root", required=True)
    parser.add_argument("--open-output", action="store_true")
    args = parser.parse_args()

    step10_root = Path(args.step10_root)
    output_root = ensure_dir(args.output_root)

    d01 = ensure_dir(output_root / "01_publication_excel")
    d02 = ensure_dir(output_root / "02_publication_ready_tsv")
    d03 = ensure_dir(output_root / "03_supporting_source_files")
    d04 = ensure_dir(output_root / "04_reports")
    d05 = ensure_dir(output_root / "05_package_zip")

    step10_summary = load_json(step10_root / "v2_step10_integrated_interpretation_package_summary.json")

    model_comparison = read_table(step10_root / "02_model_comparison" / "model_comparison_table.tsv")
    validated_treatments = read_table(step10_root / "03_validated_treatments" / "validated_treatment_table.tsv")
    recurrent_features = read_table(step10_root / "04_recurrent_spatial_features" / "recurrent_spatial_feature_table.tsv")
    recurrent_themes = read_table(step10_root / "05_recurrent_biology_themes" / "recurrent_biology_theme_table.tsv")
    figure_manifest = read_table(step10_root / "06_figures_for_presentation" / "figure_manifest.tsv")
    integration_recs = read_table(step10_root / "08_pipeline_integration_recommendations" / "pipeline_integration_recommendations.tsv")
    provenance = read_table(step10_root / "09_provenance_and_manifest" / "provenance_table.tsv")
    source_manifest = read_table(step10_root / "09_provenance_and_manifest" / "source_file_manifest.tsv")

    # Publication tables retain the columns most useful for manuscript review rather than full internal schemas.
    pub_model = short_publication_table(
        model_comparison,
        [
            "step",
            "model_branch",
            "purpose",
            "input_target",
            "feature_set",
            "split_or_validation",
            "primary_metric_name",
            "primary_metric_value",
            "secondary_metric_name",
            "secondary_metric_value",
            "validation_status",
            "notes",
        ],
    )

    pub_treatments = short_publication_table(
        validated_treatments,
        [
            "drug_key",
            "observed_test_pearson_mean",
            "observed_test_r2_mean",
            "observed_rmse_improvement_vs_baseline_mean",
            "null_test_pearson_mean_median",
            "null_test_pearson_mean_q95",
            "empirical_p_pearson",
            "fdr_q_pearson",
            "pearson_effect_vs_null_median",
            "label_shuffle_validation_status",
            "integrated_interpretation_status",
        ],
    )

    pub_features = short_publication_table(
        recurrent_features,
        [
            "feature_name",
            "feature_original",
            "biological_theme",
            "model_branch_count",
            "total_branch_score",
            "step05_registry_present",
            "step06_broad_present",
            "step07_per_treatment_present",
            "step08_curated_present",
            "step09_validated_present",
        ],
        max_rows=200,
    )

    pub_themes = short_publication_table(
        recurrent_themes,
        [
            "biological_theme",
            "model_branch_count",
            "total_branch_score",
            "step05_registry_present",
            "step06_broad_present",
            "step07_per_treatment_present",
            "step08_curated_present",
            "step09_validated_present",
            "example_features",
        ],
    )

    pub_figures = short_publication_table(
        figure_manifest,
        [
            "figure_id",
            "source_step",
            "file_name",
            "caption",
            "copied_file",
            "source_file",
        ],
    )

    pub_recs = short_publication_table(
        integration_recs,
        [
            "recommendation_id",
            "component",
            "recommendation",
            "priority",
            "rationale",
        ],
    )

    pub_provenance = short_publication_table(
        provenance,
        [
            "step",
            "root",
            "exists",
            "canonical_v1_modified",
            "v2_production_dependency_on_v1_outputs",
        ],
    )

    table_map = {
        "Model_Comparison": pub_model,
        "Validated_Treatments": pub_treatments,
        "Recurrent_Features": pub_features,
        "Biology_Themes": pub_themes,
        "Figure_Manifest": pub_figures,
        "Integration_Recs": pub_recs,
        "Provenance": pub_provenance,
    }

    pre_support_manifest = copy_supporting_files(step10_root, d03)
    table_map["Supporting_Files"] = short_publication_table(
        pre_support_manifest,
        [
            "file_type",
            "source_path",
            "copied_path",
            "size_bytes",
        ],
        max_rows=1000,
    )

    tsv_paths = {}
    for name, df in table_map.items():
        path = d02 / f"Pub_{name}.tsv"
        write_table(df, path)
        tsv_paths[name] = path

    # The Excel workbook is the main human-facing table bundle for collaborators and reviewers.
    workbook_path = d01 / "v2_integrated_publication_tables.xlsx"
    excel_success, excel_message = create_excel_workbook(table_map, workbook_path)

    if not excel_success:
        raise RuntimeError(f"Excel workbook creation failed: {excel_message}")

    sheet_summary = pd.DataFrame([
        {
            "sheet_name": name,
            "rows": int(len(df)),
            "columns": int(len(df.columns)) if not df.empty else 0,
            "description": table_description(name),
            "tsv_path": str(tsv_paths.get(name, "")),
        }
        for name, df in table_map.items()
    ])
    write_table(sheet_summary, d01 / "excel_workbook_sheet_summary.tsv")

    # Supporting source files are copied to short paths to simplify sharing and avoid path-length problems.
    support_manifest = pre_support_manifest

    tsv_manifest = pd.DataFrame([
        {
            "table_name": name,
            "path": str(path),
            "rows": int(len(table_map[name])),
            "columns": int(len(table_map[name].columns)) if not table_map[name].empty else 0,
        }
        for name, path in tsv_paths.items()
    ])
    write_table(tsv_manifest, d02 / "publication_ready_tsv_manifest.tsv")

    generated_manifest = file_manifest(output_root)
    write_table(generated_manifest, output_root / "publication_output_manifest.tsv")

    # The ZIP package is the portable publication handoff artifact.
    zip_path = d05 / "v2_publication_tables_and_supporting_files.zip"
    make_zip(output_root, zip_path)

    zip_manifest = file_manifest(output_root)
    write_table(zip_manifest, d05 / "zip_file_manifest.tsv")

    report_lines = []
    report_lines.append("V2 PUBLICATION TABLES AND SUPPORTING FILES REPORT")
    report_lines.append("")
    report_lines.append("status: pass")
    report_lines.append(f"step10_root: {step10_root}")
    report_lines.append(f"output_root: {output_root}")
    report_lines.append(f"excel_workbook: {workbook_path}")
    report_lines.append(f"zip_package: {zip_path}")
    report_lines.append(f"excel_status: {excel_message}")
    report_lines.append("")
    report_lines.append("Publication tables")
    for name, path in tsv_paths.items():
        df = table_map[name]
        report_lines.append(f"{name}: rows={len(df)} columns={len(df.columns) if not df.empty else 0} path={path}")
    report_lines.append("")
    report_lines.append("Supporting files")
    report_lines.append(f"supporting_file_count: {len(support_manifest)}")
    report_lines.append("")
    report_lines.append("Notes")
    report_lines.append("The Excel workbook contains one summary sheet and one sheet per publication table.")
    report_lines.append("Supporting source files were copied to short-path folders to avoid path-length problems.")
    report_lines.append("Copied .txt support files start with FILEPATH and SOURCE_FILEPATH headers.")

    report_path = write_text_report(d04 / "publication_tables_and_supporting_files_report.txt", "\n".join(report_lines))

    summary = {
        "status": "pass",
        "official_step": "11_make_publication_tables",
        "step10_root": str(step10_root),
        "output_root": str(output_root),
        "excel_workbook": str(workbook_path),
        "zip_package": str(zip_path),
        "n_publication_tables": int(len(table_map)),
        "n_supporting_files": int(len(support_manifest)),
        "n_generated_files": int(len(file_manifest(output_root))),
        "excel_status": excel_message,
        "production_dependency_on_v1_outputs": "no",
        "canonical_v1_scripts_modified": "no",
    }
    write_json(summary, output_root / "v2_step11_publication_tables_summary.json")

    terminal_lines = [
        "Status: pass",
        f"Step 10 root: {step10_root}",
        f"Output root: {output_root}",
        f"Excel workbook: {workbook_path}",
        f"ZIP package: {zip_path}",
        f"Report: {report_path}",
        f"Publication tables: {len(table_map)}",
        f"Supporting files: {len(support_manifest)}",
        "Production dependency on V1 outputs: no",
        "Canonical V1 scripts modified: no",
    ]

    print("")
    print(terminal_block("V2 STEP 11 PUBLICATION TABLES COMPLETE", terminal_lines))
    print("")

    print("Workbook sheet summary")
    print(sheet_summary.to_string(index=False))
    print("")

    print("Publication TSV manifest")
    print(tsv_manifest.to_string(index=False))
    print("")

    if args.open_output:
        maybe_open(output_root)
        maybe_open(d01)
        maybe_open(d02)
        maybe_open(d03)
        maybe_open(d05)

    return 0


# =============================================================================
# Script entry point
# =============================================================================

if __name__ == "__main__":
    raise SystemExit(main())
