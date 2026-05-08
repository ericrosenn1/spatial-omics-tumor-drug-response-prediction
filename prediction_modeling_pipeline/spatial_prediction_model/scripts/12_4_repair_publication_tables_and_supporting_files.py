"""
Script:
    12_N_repair_publication_tables_and_supporting_files.py

Purpose:
    Repair and complete a partially generated publication tables and supporting source files folder.

Why this exists:
    The prior publication table script could fail if the source manifest contained directory paths
    such as the project root. This repair script skips directories and only copies real files.

Design:
    New downstream repair/reporting script.
    Does not modify canonical scripts 01 to 07.
    Does not rerun models.
    Does not overwrite the final integrated package.
    Completes report, summary JSON, source manifest, README, and ZIP bundle.

Text report convention:
    Every generated .txt report starts with its own full filepath.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_txt(path: Path, body: str) -> None:
    ensure_dir(path.parent)
    path.write_text(f"FILEPATH: {path}\n\n{body}", encoding="utf-8")


def read_json(path: Path):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def read_table(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        if path.suffix.lower() == ".tsv":
            return pd.read_csv(path, sep="\t", low_memory=False)
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path, low_memory=False)
        if path.suffix.lower() == ".json":
            obj = read_json(path)
            if isinstance(obj, list):
                return pd.DataFrame(obj)
            if isinstance(obj, dict):
                return pd.DataFrame([obj])
        return pd.read_csv(path, sep=None, engine="python", low_memory=False)
    except Exception as exc:
        print(f"WARNING: failed to read table {path}: {exc}")
        return pd.DataFrame()


def is_relative_to(child: Path, parent: Path) -> bool:
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except Exception:
        return False


def safe_flat_name(path: Path, prefix: str = "") -> str:
    text = str(path)
    text = text.replace(":", "_")
    text = text.replace("\\", "__")
    text = text.replace("/", "__")
    text = "".join(ch if ch.isalnum() or ch in "._-" else "_" for ch in text)
    text = "_".join([x for x in text.split("_") if x])
    if prefix:
        text = prefix + "__" + text
    if len(text) > 180:
        digest = hashlib.sha1(str(path).encode("utf-8", errors="ignore")).hexdigest()[:12]
        suffix = path.suffix
        text = text[:140] + "__" + digest + suffix
    return text


def unique_destination(dest: Path) -> Path:
    if not dest.exists():
        return dest
    stem = dest.stem
    suffix = dest.suffix
    parent = dest.parent
    counter = 2
    while True:
        candidate = parent / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def collect_source_candidates(final_root: Path, publication_output_root: Path) -> list[Path]:
    candidates = []

    for pattern in ["*.tsv", "*.csv", "*.json", "*.txt", "*.png"]:
        for path in final_root.rglob(pattern):
            if is_relative_to(path, publication_output_root):
                continue
            candidates.append(path)

    source_manifest_path = final_root / "08_provenance_and_file_manifest" / "source_file_manifest.tsv"
    source_manifest = read_table(source_manifest_path)

    if not source_manifest.empty and "path" in source_manifest.columns:
        for raw in source_manifest["path"].dropna().astype(str).tolist():
            raw = raw.strip()
            if not raw:
                continue
            candidates.append(Path(raw))

    return candidates


def copy_supporting_sources(final_root: Path, publication_output_root: Path, max_copy_mb: float) -> pd.DataFrame:
    source_dir = ensure_dir(publication_output_root / "02_supporting_source_files")
    copied_tables_dir = ensure_dir(source_dir / "copied_source_tables_and_reports")
    copied_figures_dir = ensure_dir(source_dir / "copied_figures")

    candidates = collect_source_candidates(final_root, publication_output_root)

    rows = []
    seen = set()

    for path in candidates:
        try:
            path = Path(path)
            key = str(path.resolve()).lower() if path.exists() else str(path).lower()
        except Exception:
            key = str(path).lower()

        if key in seen:
            continue
        seen.add(key)

        row = {
            "source_path": str(path),
            "exists": False,
            "is_file": False,
            "is_directory": False,
            "copied": False,
            "reason": "",
            "size_mb": "",
            "destination_path": "",
        }

        try:
            if not path.exists():
                row["reason"] = "missing"
                rows.append(row)
                continue

            row["exists"] = True
            row["is_file"] = path.is_file()
            row["is_directory"] = path.is_dir()

            if path.is_dir():
                row["reason"] = "skipped_directory"
                rows.append(row)
                continue

            if not path.is_file():
                row["reason"] = "skipped_not_regular_file"
                rows.append(row)
                continue

            size_mb = path.stat().st_size / (1024 * 1024)
            row["size_mb"] = round(size_mb, 4)

            if size_mb > max_copy_mb:
                row["reason"] = f"skipped_larger_than_{max_copy_mb}_MB"
                rows.append(row)
                continue

            if is_relative_to(path, publication_output_root):
                row["reason"] = "skipped_publication_output_file"
                rows.append(row)
                continue

            suffix = path.suffix.lower()
            dest_base = copied_figures_dir if suffix in {".png", ".jpg", ".jpeg", ".pdf"} else copied_tables_dir

            try:
                rel = path.relative_to(final_root)
                dest_name = safe_flat_name(rel)
            except Exception:
                dest_name = safe_flat_name(path, prefix="external")

            dest = unique_destination(dest_base / dest_name)

            try:
                shutil.copy2(path, dest)
                row["copied"] = True
                row["reason"] = "copied"
                row["destination_path"] = str(dest)
            except Exception as exc:
                row["reason"] = "copy_failed: " + str(exc)

            rows.append(row)

        except Exception as exc:
            row["reason"] = "unexpected_error: " + str(exc)
            rows.append(row)

    manifest = pd.DataFrame(rows)
    if not manifest.empty:
        manifest = manifest.sort_values(["copied", "reason", "source_path"], ascending=[False, True, True])

    manifest_path = source_dir / "supporting_source_file_manifest.tsv"
    manifest.to_csv(manifest_path, sep="\t", index=False)

    readme_lines = []
    readme_lines.append("PUBLICATION SUPPORTING SOURCE FILES")
    readme_lines.append("")
    readme_lines.append(f"Final integrated package root: {final_root}")
    readme_lines.append(f"Publication output root: {publication_output_root}")
    readme_lines.append("")
    readme_lines.append("This folder contains copied source tables, reports, and figures used by the publication table package.")
    readme_lines.append("Directories listed in source manifests are intentionally skipped.")
    readme_lines.append("Large files are intentionally skipped to keep the source bundle portable.")
    readme_lines.append("")
    readme_lines.append(f"Supporting source manifest: {manifest_path}")

    write_txt(source_dir / "README_supporting_source_files.txt", "\n".join(readme_lines))

    return manifest


def build_publication_tsv_manifest(publication_output_root: Path) -> pd.DataFrame:
    tsv_dir = publication_output_root / "03_publication_ready_tsv"
    rows = []

    if tsv_dir.exists():
        for path in sorted(tsv_dir.glob("*.tsv")):
            if path.name == "publication_ready_tsv_manifest.tsv":
                continue

            rows.append({
                "table_name": path.stem,
                "path": str(path),
                "size_kb": round(path.stat().st_size / 1024, 3),
                "last_modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            })

    manifest = pd.DataFrame(rows)
    manifest_path = tsv_dir / "publication_ready_tsv_manifest.tsv"
    ensure_dir(tsv_dir)
    manifest.to_csv(manifest_path, sep="\t", index=False)
    return manifest


def get_workbook_sheet_summary(publication_output_root: Path) -> pd.DataFrame:
    summary_path = publication_output_root / "01_publication_excel" / "excel_workbook_sheet_summary.tsv"

    if summary_path.exists():
        return read_table(summary_path)

    workbook_path = publication_output_root / "01_publication_excel" / "final_integrated_publication_tables.xlsx"

    rows = []
    if workbook_path.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(workbook_path, read_only=True, data_only=False)
            for ws in wb.worksheets:
                rows.append({
                    "Sheet name": ws.title,
                    "Type": "Workbook sheet",
                    "Description": "",
                    "Rows": ws.max_row,
                    "Columns": ws.max_column,
                    "Source path": str(workbook_path),
                })
            wb.close()
        except Exception as exc:
            rows.append({
                "Sheet name": "unknown",
                "Type": "Workbook sheet scan failed",
                "Description": str(exc),
                "Rows": "",
                "Columns": "",
                "Source path": str(workbook_path),
            })

    summary = pd.DataFrame(rows)
    ensure_dir(summary_path.parent)
    summary.to_csv(summary_path, sep="\t", index=False)
    return summary


def create_zip(publication_output_root: Path) -> tuple[Path, pd.DataFrame]:
    zip_path = publication_output_root / f"{publication_output_root.name}.zip"

    if zip_path.exists():
        zip_path.unlink()

    rows = []

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in publication_output_root.rglob("*"):
            if path == zip_path:
                continue

            if not path.is_file():
                continue

            try:
                zf.write(path, path.relative_to(publication_output_root))
                rows.append({
                    "path": str(path),
                    "included_in_zip": True,
                    "reason": "included",
                    "size_mb": round(path.stat().st_size / (1024 * 1024), 4),
                })
            except Exception as exc:
                rows.append({
                    "path": str(path),
                    "included_in_zip": False,
                    "reason": str(exc),
                    "size_mb": "",
                })

    zip_manifest = pd.DataFrame(rows)
    zip_manifest_path = publication_output_root / "02_supporting_source_files" / "zip_file_manifest.tsv"
    ensure_dir(zip_manifest_path.parent)
    zip_manifest.to_csv(zip_manifest_path, sep="\t", index=False)

    return zip_path, zip_manifest


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--final-run-root", required=True)
    parser.add_argument("--publication-output-root", required=True)
    parser.add_argument("--max-copy-mb", type=float, default=50.0)
    args = parser.parse_args()

    final_root = Path(args.final_run_root)
    publication_output_root = Path(args.publication_output_root)

    if not final_root.exists():
        raise FileNotFoundError(f"Final integrated package root not found: {final_root}")

    if not publication_output_root.exists():
        raise FileNotFoundError(f"Publication output root not found: {publication_output_root}")

    report_dir = ensure_dir(publication_output_root / "04_reports")
    excel_dir = ensure_dir(publication_output_root / "01_publication_excel")
    source_dir = ensure_dir(publication_output_root / "02_supporting_source_files")
    tsv_dir = ensure_dir(publication_output_root / "03_publication_ready_tsv")

    workbook_path = excel_dir / "final_integrated_publication_tables.xlsx"
    final_run_summary_path = final_root / "01_model_milestone_summary" / "run_summary.json"
    final_run_summary = read_json(final_run_summary_path) or {}

    tsv_manifest = build_publication_tsv_manifest(publication_output_root)
    workbook_sheet_summary = get_workbook_sheet_summary(publication_output_root)
    source_manifest = copy_supporting_sources(final_root, publication_output_root, max_copy_mb=args.max_copy_mb)

    copied_count = int(source_manifest["copied"].astype(bool).sum()) if not source_manifest.empty else 0
    skipped_count = int((~source_manifest["copied"].astype(bool)).sum()) if not source_manifest.empty else 0

    report_path = report_dir / "publication_tables_and_supporting_files_report.txt"
    summary_json_path = publication_output_root / "publication_tables_and_supporting_files_summary.json"

    report_lines = []
    report_lines.append("PUBLICATION TABLES AND SUPPORTING SOURCE FILES REPAIR REPORT")
    report_lines.append("")
    report_lines.append("Status: repaired and completed")
    report_lines.append("")
    report_lines.append(f"Final integrated package root: {final_root}")
    report_lines.append(f"Publication output root: {publication_output_root}")
    report_lines.append("")
    report_lines.append("Reason for repair")
    report_lines.append("The prior script failed while copying supporting source files because a directory path was treated as a file.")
    report_lines.append("This repair script skips directories, missing paths, oversized files, and inaccessible paths.")
    report_lines.append("")
    report_lines.append("Final integrated package summary")
    report_lines.append(f"Model families summarized: {final_run_summary.get('model_families_summarized', '')}")
    report_lines.append(f"Validated treatment models: {final_run_summary.get('validated_treatment_models', '')}")
    report_lines.append(f"Validated treatment models passing FDR q <= 0.10: {final_run_summary.get('validated_treatment_models_passing_fdr_0_10', '')}")
    report_lines.append(f"Recurrent spatial features: {final_run_summary.get('recurrent_spatial_features', '')}")
    report_lines.append(f"Recurrent biology themes: {final_run_summary.get('recurrent_biology_themes', '')}")
    report_lines.append("")
    report_lines.append("Generated publication outputs")
    report_lines.append(f"Combined Excel workbook exists: {workbook_path.exists()}")
    report_lines.append(f"Combined Excel workbook: {workbook_path}")
    report_lines.append(f"Publication ready TSV folder: {tsv_dir}")
    report_lines.append(f"Supporting source folder: {source_dir}")
    report_lines.append("")
    report_lines.append("Workbook sheet summary")
    if workbook_sheet_summary.empty:
        report_lines.append("No workbook sheet summary rows were available.")
    else:
        for _, row in workbook_sheet_summary.iterrows():
            report_lines.append(
                f"{row.get('Sheet name', '')}: type={row.get('Type', '')}, rows={row.get('Rows', '')}, columns={row.get('Columns', '')}"
            )
    report_lines.append("")
    report_lines.append("Publication TSV summary")
    if tsv_manifest.empty:
        report_lines.append("No publication ready TSV rows were found.")
    else:
        for _, row in tsv_manifest.iterrows():
            report_lines.append(f"{row.get('table_name', '')}: {row.get('path', '')}")
    report_lines.append("")
    report_lines.append("Supporting source copy summary")
    report_lines.append(f"Source files copied: {copied_count}")
    report_lines.append(f"Source files skipped, missing, directory, or inaccessible: {skipped_count}")
    report_lines.append(f"Supporting source manifest: {source_dir / 'supporting_source_file_manifest.tsv'}")

    write_txt(report_path, "\n".join(report_lines))

    summary = {
        "status": "repaired_and_completed",
        "final_integrated_package_root": str(final_root),
        "publication_output_root": str(publication_output_root),
        "combined_excel_workbook": str(workbook_path),
        "combined_excel_workbook_exists": workbook_path.exists(),
        "publication_ready_tsv_folder": str(tsv_dir),
        "supporting_source_folder": str(source_dir),
        "report": str(report_path),
        "workbook_sheets": int(len(workbook_sheet_summary)),
        "publication_ready_tsv_files": int(len(tsv_manifest)),
        "source_files_copied": copied_count,
        "source_files_skipped_or_missing_or_inaccessible": skipped_count,
        "model_families_summarized": final_run_summary.get("model_families_summarized", None),
        "validated_treatment_models": final_run_summary.get("validated_treatment_models", None),
        "validated_treatment_models_passing_fdr_0_10": final_run_summary.get("validated_treatment_models_passing_fdr_0_10", None),
    }

    summary_json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    zip_path, zip_manifest = create_zip(publication_output_root)

    summary["supporting_zip_bundle"] = str(zip_path)
    summary["zip_manifest"] = str(source_dir / "zip_file_manifest.tsv")
    summary["zip_files_included"] = int(zip_manifest["included_in_zip"].astype(bool).sum()) if not zip_manifest.empty else 0
    summary_json_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print("")
    print("=" * 100)
    print("PUBLICATION TABLES AND SUPPORTING FILES REPAIR COMPLETE")
    print("=" * 100)
    print("Final integrated package root:", final_root)
    print("Publication output root:", publication_output_root)
    print("Report:", report_path)
    print("Summary JSON:", summary_json_path)
    print("Combined Excel workbook:", workbook_path)
    print("Publication ready TSV folder:", tsv_dir)
    print("Supporting source folder:", source_dir)
    print("Supporting ZIP bundle:", zip_path)
    print("Workbook sheets:", summary["workbook_sheets"])
    print("Publication ready TSV files:", summary["publication_ready_tsv_files"])
    print("Source files copied:", summary["source_files_copied"])
    print("Source files skipped, missing, directory, or inaccessible:", summary["source_files_skipped_or_missing_or_inaccessible"])
    print("ZIP files included:", summary["zip_files_included"])
    print("")
    print("Report first line should start with FILEPATH.")
    print("")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
