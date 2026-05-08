"""
Script:
    12_N_make_publication_tables_and_supporting_files.py

Purpose:
    Build publication ready Excel tables and supporting source file bundle from a completed
    final integrated spatial response interpretation package.

Design:
    New downstream reporting script.
    Does not modify canonical scripts 01 to 07.
    Does not retrain models.
    Reads the latest completed final integrated package.
    Writes a new organized publication table and source bundle folder.

Text report convention:
    Every generated .txt report starts with its own full filepath.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except Exception as exc:
    raise ImportError(
        "openpyxl is required to create the publication Excel workbook. "
        "Install it in the active environment with: python -m pip install openpyxl"
    ) from exc


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def write_txt(path: Path, body: str) -> None:
    ensure_dir(path.parent)
    path.write_text(f"FILEPATH: {path}\n\n{body}", encoding="utf-8")


def read_json(path: Path):
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def read_table(path: Path) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame()
    try:
        if path.suffix.lower() == ".tsv":
            return pd.read_csv(path, sep="\t", low_memory=False)
        if path.suffix.lower() == ".csv":
            return pd.read_csv(path, low_memory=False)
        if path.suffix.lower() == ".json":
            obj = read_json(path)
            if isinstance(obj, list):
                return pd.DataFrame(obj)
            if isinstance(obj, dict):
                return pd.DataFrame([obj])
        return pd.read_csv(path, sep=None, engine="python", low_memory=False)
    except Exception as exc:
        print(f"WARNING: failed to read {path}: {exc}")
        return pd.DataFrame()


def find_col(df: pd.DataFrame, candidates: List[str]) -> str | None:
    if df is None or df.empty:
        return None
    exact = {str(c).lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in exact:
            return exact[c.lower()]
    for c in candidates:
        needle = c.lower()
        for col in df.columns:
            if needle in str(col).lower():
                return col
    return None


def numeric_col(df: pd.DataFrame, col: str | None) -> pd.Series:
    if col is None or col not in df.columns:
        return pd.Series([np.nan] * len(df), index=df.index)
    return pd.to_numeric(df[col], errors="coerce")


def bool_to_yes_no(x) -> str:
    s = str(x).strip().lower()
    if s in {"true", "1", "yes", "y"}:
        return "Yes"
    if s in {"false", "0", "no", "n"}:
        return "No"
    if s in {"nan", "none", ""}:
        return ""
    return str(x)


def human_label(x) -> str:
    x = str(x)
    x = x.replace("_", " ")
    x = re.sub(r"\s+", " ", x).strip()
    return x


def round_float(x, digits=4):
    try:
        if pd.isna(x):
            return ""
        return round(float(x), digits)
    except Exception:
        return x


def add_rank(df: pd.DataFrame, sort_col: str | None = None, ascending: bool = False) -> pd.DataFrame:
    out = df.copy()
    if sort_col and sort_col in out.columns:
        out[sort_col] = pd.to_numeric(out[sort_col], errors="coerce")
        out = out.sort_values(sort_col, ascending=ascending, na_position="last").reset_index(drop=True)
    out.insert(0, "Rank", range(1, len(out) + 1))
    return out


def clean_for_excel_value(value):
    if isinstance(value, (list, dict)):
        return json.dumps(value)
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if pd.isna(value):
        return None
    return value


def safe_sheet_name(name: str, used: set) -> str:
    base = re.sub(r"[\[\]\*\?\/\\:]", "_", str(name)).strip()
    base = base[:31] if len(base) > 31 else base
    if not base:
        base = "Sheet"
    candidate = base
    idx = 2
    while candidate in used:
        suffix = f"_{idx}"
        candidate = base[: 31 - len(suffix)] + suffix
        idx += 1
    used.add(candidate)
    return candidate


def safe_table_name(name: str, used: set) -> str:
    base = re.sub(r"[^A-Za-z0-9_]", "_", str(name))
    base = re.sub(r"_+", "_", base).strip("_")
    if not base:
        base = "Table"
    if not re.match(r"^[A-Za-z_]", base):
        base = "T_" + base
    base = base[:200]
    candidate = base
    idx = 2
    while candidate in used:
        candidate = f"{base[:190]}_{idx}"
        idx += 1
    used.add(candidate)
    return candidate


def style_sheet(ws, data_start_row: int, data_start_col: int, n_rows: int, n_cols: int, table_name: str | None = None):
    title_fill = PatternFill("solid", fgColor="1F4E78")
    subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A2"].fill = subtitle_fill
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = ws.cell(row=data_start_row + 1, column=data_start_col)

    for col_idx in range(data_start_col, data_start_col + n_cols):
        cell = ws.cell(row=data_start_row, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        cell.border = border

    for row in ws.iter_rows(min_row=data_start_row, max_row=max(data_start_row, data_start_row + n_rows), min_col=data_start_col, max_col=data_start_col + n_cols - 1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 8
        for cell in ws[letter]:
            val = cell.value
            if val is None:
                continue
            for piece in str(val).split("\n"):
                max_len = max(max_len, min(len(piece), 80))
        width = min(max(max_len + 2, 10), 48)
        ws.column_dimensions[letter].width = width

    for row_idx in range(1, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 18

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 30

    if table_name and n_cols > 0 and n_rows > 0:
        ref = f"{get_column_letter(data_start_col)}{data_start_row}:{get_column_letter(data_start_col + n_cols - 1)}{data_start_row + n_rows}"
        tab = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        try:
            ws.add_table(tab)
        except Exception:
            pass


def write_df_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame, description: str, source_path: str, used_sheets: set, used_tables: set) -> str:
    ws_name = safe_sheet_name(sheet_name, used_sheets)
    ws = wb.create_sheet(ws_name)

    if df is None:
        df = pd.DataFrame()

    ws["A1"] = ws_name
    ws["A2"] = description
    ws["A3"] = f"Source: {source_path}" if source_path else "Source: generated from integrated package"

    start_row = 5
    start_col = 1

    if df.empty:
        ws.cell(row=start_row, column=start_col, value="No rows available")
        style_sheet(ws, start_row, start_col, 1, 1, None)
        return ws_name

    out = df.copy()

    for col_idx, col_name in enumerate(out.columns, start=start_col):
        ws.cell(row=start_row, column=col_idx, value=str(col_name))

    for row_idx, (_, row) in enumerate(out.iterrows(), start=start_row + 1):
        for col_idx, col_name in enumerate(out.columns, start=start_col):
            ws.cell(row=row_idx, column=col_idx, value=clean_for_excel_value(row[col_name]))

    table_name = safe_table_name(ws_name, used_tables)
    style_sheet(ws, start_row, start_col, len(out), len(out.columns), table_name)
    return ws_name


def make_publication_model_comparison(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = pd.DataFrame()
    out["Model family"] = df.get("model_family", "").map(human_label)
    out["Model name"] = df.get("model_name", "")
    out["Purpose"] = df.get("purpose", "")
    out["Unit of prediction"] = df.get("unit_of_prediction", "").map(human_label)
    out["Target"] = df.get("target", "")
    out["Inputs"] = df.get("inputs", "")
    out["Validation type"] = df.get("validation_type", "")
    out["Primary metric"] = df.get("primary_metric_name", "").map(human_label)
    out["Primary value"] = df.get("primary_metric_value", "").map(lambda x: round_float(x, 4))
    out["Secondary metric"] = df.get("secondary_metric_name", "").map(human_label)
    out["Secondary value"] = df.get("secondary_metric_value", "").map(lambda x: round_float(x, 4))
    out["Spatial feature fraction"] = df.get("spatial_feature_fraction", "").map(lambda x: round_float(x, 4))
    out["Recommended use"] = df.get("recommended_use", "")
    out["Interpretation note"] = df.get("notes", "")
    return out


def make_publication_validated_treatments(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    drug_col = find_col(df, ["drug_label", "treatment", "treatment_name", "drug_key"])
    obs_col = find_col(df, ["observed_test_pearson_mean", "observed_mean_test_pearson"])
    r2_col = find_col(df, ["observed_test_r2_mean", "observed_mean_test_r2"])
    p_col = find_col(df, ["empirical_p_value", "empirical_p"])
    fdr_col = find_col(df, ["fdr_q_value", "bh_fdr_q_value", "q_value"])
    null_col = find_col(df, ["null_test_pearson_mean", "null_mean_test_pearson"])
    null_q95_col = find_col(df, ["null_test_pearson_q95", "null_q95"])
    pass_col = find_col(df, ["passes_fdr10", "passes_fdr_10"])
    n_col = find_col(df, ["n_samples"])

    out = pd.DataFrame()
    out["Treatment"] = df[drug_col].astype(str) if drug_col else df.index.astype(str)
    if n_col:
        out["Samples"] = numeric_col(df, n_col).astype("Int64")
    if obs_col:
        out["Observed mean test Pearson"] = numeric_col(df, obs_col).map(lambda x: round_float(x, 4))
    if r2_col:
        out["Observed mean test R2"] = numeric_col(df, r2_col).map(lambda x: round_float(x, 4))
    if null_col:
        out["Null mean test Pearson"] = numeric_col(df, null_col).map(lambda x: round_float(x, 4))
    if null_q95_col:
        out["Null 95th percentile Pearson"] = numeric_col(df, null_q95_col).map(lambda x: round_float(x, 4))
    if p_col:
        out["Empirical P value"] = numeric_col(df, p_col).map(lambda x: round_float(x, 5))
    if fdr_col:
        out["FDR q value"] = numeric_col(df, fdr_col).map(lambda x: round_float(x, 5))
    if pass_col:
        out["Passes FDR 0.10"] = df[pass_col].map(bool_to_yes_no)
    if obs_col:
        out = add_rank(out, "Observed mean test Pearson", ascending=False)
    else:
        out = add_rank(out)
    return out


def make_publication_broad_targets(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    target_col = find_col(df, ["target_name", "target"])
    out = pd.DataFrame()
    out["Target"] = df[target_col].astype(str) if target_col else df.index.astype(str)

    for source, label in [
        ("test_pearson_mean", "Mean test Pearson"),
        ("test_pearson_median", "Median test Pearson"),
        ("test_pearson_q025", "Pearson lower 2.5 percent"),
        ("test_pearson_q975", "Pearson upper 97.5 percent"),
        ("test_r2_mean", "Mean test R2"),
        ("test_mae_mean", "Mean test MAE"),
        ("rmse_improvement_vs_baseline_mean", "Mean RMSE improvement versus baseline"),
    ]:
        col = find_col(df, [source, label])
        if col:
            out[label] = numeric_col(df, col).map(lambda x: round_float(x, 4))

    if "Mean test Pearson" in out.columns:
        out = add_rank(out, "Mean test Pearson", ascending=False)
    else:
        out = add_rank(out)

    return out


def make_publication_recurrent_features(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    feature_col = find_col(df, ["feature", "feature_name", "feature_original"])
    out = pd.DataFrame()
    out["Feature"] = df[feature_col].astype(str) if feature_col else df.index.astype(str)

    for source, label in [
        ("source_count", "Model branch count"),
        ("source_branches", "Model branches"),
        ("biological_themes", "Biology themes"),
        ("mean_score", "Mean contribution score"),
        ("max_score", "Maximum contribution score"),
    ]:
        col = find_col(df, [source, label])
        if col:
            if "score" in source:
                out[label] = numeric_col(df, col).map(lambda x: round_float(x, 5))
            else:
                out[label] = df[col]

    if "Model branch count" in out.columns:
        out = add_rank(out, "Model branch count", ascending=False)
    else:
        out = add_rank(out)

    return out


def make_publication_biology_themes(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    theme_col = find_col(df, ["theme", "biological_theme"])
    out = pd.DataFrame()
    out["Biology theme"] = df[theme_col].astype(str) if theme_col else df.index.astype(str)

    for source, label in [
        ("source_count", "Model branch count"),
        ("source_branches", "Model branches"),
        ("mean_score", "Mean contribution score"),
        ("max_score", "Maximum contribution score"),
    ]:
        col = find_col(df, [source, label])
        if col:
            if "score" in source:
                out[label] = numeric_col(df, col).map(lambda x: round_float(x, 5))
            else:
                out[label] = df[col]

    if "Model branch count" in out.columns:
        out = add_rank(out, "Model branch count", ascending=False)
    else:
        out = add_rank(out)

    return out


def make_publication_figure_manifest(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = pd.DataFrame()
    for source, label in [
        ("figure_id", "Figure or table ID"),
        ("title", "Title"),
        ("description", "Description"),
        ("recommended_use", "Recommended use"),
        ("path", "Path"),
    ]:
        col = find_col(df, [source, label])
        if col:
            out[label] = df[col]
    return out


def make_publication_integration_recs(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df

    out = pd.DataFrame()
    for source, label in [
        ("pipeline_component", "Pipeline component"),
        ("current_status", "Current status"),
        ("recommended_decision", "Recommended decision"),
        ("promote_to_canonical", "Promote to canonical"),
        ("rationale", "Rationale"),
        ("notes", "Notes"),
    ]:
        col = find_col(df, [source, label])
        if col:
            out[label] = df[col].map(human_label) if source in {"pipeline_component", "current_status", "recommended_decision"} else df[col]
    return out


def write_publication_tsvs(pub_tables: Dict[str, Tuple[pd.DataFrame, str]], out_dir: Path) -> List[dict]:
    ensure_dir(out_dir)
    rows = []
    for name, (df, description) in pub_tables.items():
        path = out_dir / f"{name}.tsv"
        df.to_csv(path, sep="\t", index=False)
        rows.append({
            "table_name": name,
            "description": description,
            "rows": int(len(df)),
            "columns": int(len(df.columns)),
            "path": str(path),
        })
    return rows


def copy_supporting_sources(final_root: Path, output_root: Path, source_manifest: pd.DataFrame, max_copy_mb: float = 50.0) -> pd.DataFrame:
    source_dir = ensure_dir(output_root / "02_supporting_source_files")
    copied_dir = ensure_dir(source_dir / "copied_source_files")
    figure_dir = ensure_dir(source_dir / "copied_figures")

    candidates = []

    for pattern in ["*.tsv", "*.csv", "*.json", "*.txt", "*.png"]:
        for p in final_root.rglob(pattern):
            if output_root in p.parents:
                continue
            candidates.append(p)

    if source_manifest is not None and not source_manifest.empty and "path" in source_manifest.columns:
        for item in source_manifest["path"].dropna().astype(str).tolist():
            p = Path(item)
            if p.exists():
                candidates.append(p)

    seen = set()
    rows = []

    for path in candidates:
        path = Path(path)
        key = str(path.resolve()).lower()
        if key in seen:
            continue
        seen.add(key)

        exists = path.exists()
        size_mb = path.stat().st_size / (1024 * 1024) if exists else np.nan

        if not exists:
            rows.append({
                "source_path": str(path),
                "copied": False,
                "reason": "missing",
                "size_mb": "",
                "destination_path": "",
            })
            continue

        if size_mb > max_copy_mb:
            rows.append({
                "source_path": str(path),
                "copied": False,
                "reason": f"skipped because file is larger than {max_copy_mb} MB",
                "size_mb": round(size_mb, 3),
                "destination_path": "",
            })
            continue

        rel = None
        try:
            rel = path.relative_to(final_root)
        except Exception:
            try:
                rel = path.relative_to(final_root.parent.parent)
            except Exception:
                rel = Path(path.name)

        rel_text = str(rel).replace(":", "_").replace("\\", "__").replace("/", "__")
        dest_base = figure_dir if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".pdf"} else copied_dir
        dest = dest_base / rel_text

        counter = 2
        while dest.exists():
            dest = dest_base / f"{dest.stem}_{counter}{dest.suffix}"
            counter += 1

        shutil.copy2(path, dest)

        rows.append({
            "source_path": str(path),
            "copied": True,
            "reason": "copied",
            "size_mb": round(size_mb, 3),
            "destination_path": str(dest),
        })

    manifest = pd.DataFrame(rows).sort_values(["copied", "source_path"], ascending=[False, True])
    manifest_path = source_dir / "supporting_source_file_manifest.tsv"
    manifest.to_csv(manifest_path, sep="\t", index=False)

    readme = []
    readme.append("PUBLICATION SUPPORTING SOURCE FILES")
    readme.append("")
    readme.append(f"Final integrated package root: {final_root}")
    readme.append(f"Publication/source bundle root: {output_root}")
    readme.append("")
    readme.append("This folder contains copied source tables, reports, figures, publication ready TSV tables, and the combined Excel workbook.")
    readme.append("Large raw modeling files may be skipped intentionally to keep this bundle portable.")
    readme.append("")
    readme.append(f"Supporting manifest: {manifest_path}")

    write_txt(source_dir / "README_supporting_source_files.txt", "\n".join(readme))

    return manifest


def create_zip(output_root: Path, workbook_path: Path) -> Path:
    zip_path = output_root / f"{output_root.name}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for p in output_root.rglob("*"):
            if p == zip_path:
                continue
            if p.is_file():
                zf.write(p, p.relative_to(output_root))
        if workbook_path.exists() and workbook_path.parent not in output_root.parents:
            zf.write(workbook_path, workbook_path.name)
    return zip_path


def build_workbook(
    workbook_path: Path,
    run_summary: dict,
    publication_tables: Dict[str, Tuple[pd.DataFrame, str, str]],
    raw_tables: Dict[str, Tuple[pd.DataFrame, str, str]],
) -> pd.DataFrame:
    ensure_dir(workbook_path.parent)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    used_sheets = {"Summary"}
    used_tables = set()

    title_fill = PatternFill("solid", fgColor="1F4E78")
    header_fill = PatternFill("solid", fgColor="5B9BD5")
    subtitle_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Final integrated publication tables workbook"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].fill = subtitle_fill
    ws["A3"] = f"Output root: {run_summary.get('output_root', '')}"

    summary_items = [
        ("Model families summarized", run_summary.get("model_families_summarized", "")),
        ("Validated treatment models", run_summary.get("validated_treatment_models", "")),
        ("Validated treatment models passing FDR q <= 0.10", run_summary.get("validated_treatment_models_passing_fdr_0_10", "")),
        ("Recurrent spatial features", run_summary.get("recurrent_spatial_features", "")),
        ("Recurrent biology themes", run_summary.get("recurrent_biology_themes", "")),
        ("Figures and main tables", run_summary.get("figures_and_main_tables", "")),
    ]

    ws["A5"] = "Package summary"
    ws["A5"].font = Font(bold=True)
    for i, (k, v) in enumerate(summary_items, start=6):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)

    sheet_summary = []
    row_start = 15
    headers = ["Sheet name", "Type", "Description", "Rows", "Columns", "Source path"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row_start, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border

    sheet_summary.append({
        "Sheet name": "Summary",
        "Type": "Workbook index",
        "Description": "Workbook summary and sheet index",
        "Rows": "",
        "Columns": "",
        "Source path": str(workbook_path),
    })

    for sheet_name, (df, description, source) in publication_tables.items():
        actual_name = write_df_sheet(wb, sheet_name, df, description, source, used_sheets, used_tables)
        sheet_summary.append({
            "Sheet name": actual_name,
            "Type": "Publication ready table",
            "Description": description,
            "Rows": int(len(df)),
            "Columns": int(len(df.columns)),
            "Source path": source,
        })

    for sheet_name, (df, description, source) in raw_tables.items():
        actual_name = write_df_sheet(wb, sheet_name, df, description, source, used_sheets, used_tables)
        sheet_summary.append({
            "Sheet name": actual_name,
            "Type": "Raw source table",
            "Description": description,
            "Rows": int(len(df)),
            "Columns": int(len(df.columns)),
            "Source path": source,
        })

    for i, row in enumerate(sheet_summary, start=row_start + 1):
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(i, c, row[h])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 28 if col_idx not in {3, 6} else 55

    ws.freeze_panes = "A16"
    ws.auto_filter.ref = f"A{row_start}:F{row_start + len(sheet_summary)}"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(workbook_path)

    return pd.DataFrame(sheet_summary)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--final-run-root", required=True)
    parser.add_argument("--output-root", required=True)
    parser.add_argument("--max-copy-mb", type=float, default=50.0)
    args = parser.parse_args()

    final_root = Path(args.final_run_root)
    output_root = ensure_dir(Path(args.output_root))

    excel_dir = ensure_dir(output_root / "01_publication_excel")
    source_dir = ensure_dir(output_root / "02_supporting_source_files")
    tsv_dir = ensure_dir(output_root / "03_publication_ready_tsv")
    reports_dir = ensure_dir(output_root / "04_reports")

    paths = {
        "run_summary": final_root / "01_model_milestone_summary" / "run_summary.json",
        "master_summary": final_root / "01_model_milestone_summary" / "master_summary_report.txt",
        "model_comparison": final_root / "01_model_milestone_summary" / "model_comparison_table.tsv",
        "broad_targets": final_root / "01_model_milestone_summary" / "broad_residual_target_comparison.tsv",
        "validated_treatments": final_root / "02_validated_treatment_models" / "validated_treatment_table.tsv",
        "recurrent_features": final_root / "03_recurrent_spatial_features" / "recurrent_spatial_feature_table.tsv",
        "recurrent_features_long": final_root / "03_recurrent_spatial_features" / "recurrent_spatial_feature_long.tsv",
        "biology_themes": final_root / "04_recurrent_biology_themes" / "recurrent_biology_theme_table.tsv",
        "biology_themes_long": final_root / "04_recurrent_biology_themes" / "recurrent_biology_theme_long.tsv",
        "figure_manifest": final_root / "05_figures_for_presentation" / "figure_manifest.tsv",
        "figure_captions": final_root / "05_figures_for_presentation" / "figure_captions.txt",
        "publication_model_design": final_root / "05_figures_for_presentation" / "table01_publication_style_model_comparison.tsv",
        "methods_results_discussion": final_root / "06_methods_results_discussion_report" / "methods_results_discussion_narrative.txt",
        "integration_recs": final_root / "07_pipeline_integration_recommendations" / "pipeline_integration_recommendations.tsv",
        "integration_recs_text": final_root / "07_pipeline_integration_recommendations" / "pipeline_integration_recommendations.txt",
        "source_manifest": final_root / "08_provenance_and_file_manifest" / "source_file_manifest.tsv",
        "provenance": final_root / "08_provenance_and_file_manifest" / "provenance_table.tsv",
        "generated_manifest": final_root / "08_provenance_and_file_manifest" / "generated_output_manifest.tsv",
    }

    run_summary = read_json(paths["run_summary"]) or {}

    raw_tables = {
        "Raw_Model_Comparison": (read_table(paths["model_comparison"]), "Raw integrated model comparison table", str(paths["model_comparison"])),
        "Raw_Validated_Treat": (read_table(paths["validated_treatments"]), "Raw validated treatment label shuffle table", str(paths["validated_treatments"])),
        "Raw_Broad_Targets": (read_table(paths["broad_targets"]), "Raw broad residual target comparison table", str(paths["broad_targets"])),
        "Raw_Recurrent_Features": (read_table(paths["recurrent_features"]), "Raw recurrent spatial feature table", str(paths["recurrent_features"])),
        "Raw_Biology_Themes": (read_table(paths["biology_themes"]), "Raw recurrent biology theme table", str(paths["biology_themes"])),
        "Raw_Figure_Manifest": (read_table(paths["figure_manifest"]), "Raw final figure manifest", str(paths["figure_manifest"])),
        "Raw_Integration_Recs": (read_table(paths["integration_recs"]), "Raw canonical integration recommendation table", str(paths["integration_recs"])),
        "Raw_Source_Manifest": (read_table(paths["source_manifest"]), "Raw source file manifest", str(paths["source_manifest"])),
        "Raw_Provenance": (read_table(paths["provenance"]), "Raw provenance table", str(paths["provenance"])),
        "Raw_Generated_Manifest": (read_table(paths["generated_manifest"]), "Raw generated output manifest", str(paths["generated_manifest"])),
    }

    model_comparison_raw = raw_tables["Raw_Model_Comparison"][0]
    validated_raw = raw_tables["Raw_Validated_Treat"][0]
    broad_raw = raw_tables["Raw_Broad_Targets"][0]
    features_raw = raw_tables["Raw_Recurrent_Features"][0]
    themes_raw = raw_tables["Raw_Biology_Themes"][0]
    figure_raw = raw_tables["Raw_Figure_Manifest"][0]
    recs_raw = raw_tables["Raw_Integration_Recs"][0]

    publication_tables = {
        "Pub_Model_Comparison": (
            make_publication_model_comparison(model_comparison_raw),
            "Publication ready model comparison table with model purpose, target, inputs, validation, metrics, and recommended use",
            str(paths["model_comparison"]),
        ),
        "Pub_Validated_Treatments": (
            make_publication_validated_treatments(validated_raw),
            "Publication ready validated Tier 1 treatment table from label shuffle validation",
            str(paths["validated_treatments"]),
        ),
        "Pub_Broad_Targets": (
            make_publication_broad_targets(broad_raw),
            "Publication ready broad residual target comparison table",
            str(paths["broad_targets"]),
        ),
        "Pub_Recurrent_Features": (
            make_publication_recurrent_features(features_raw),
            "Publication ready recurrent spatial feature table across model branches",
            str(paths["recurrent_features"]),
        ),
        "Pub_Biology_Themes": (
            make_publication_biology_themes(themes_raw),
            "Publication ready recurrent biology theme table across model branches",
            str(paths["biology_themes"]),
        ),
        "Pub_Figure_Manifest": (
            make_publication_figure_manifest(figure_raw),
            "Publication ready figure and table manifest",
            str(paths["figure_manifest"]),
        ),
        "Pub_Integration_Recs": (
            make_publication_integration_recs(recs_raw),
            "Publication ready canonical integration recommendation table",
            str(paths["integration_recs"]),
        ),
    }

    publication_tsv_rows = write_publication_tsvs(
        {k: (v[0], v[1]) for k, v in publication_tables.items()},
        tsv_dir,
    )

    workbook_path = excel_dir / "final_integrated_publication_tables.xlsx"
    sheet_summary = build_workbook(workbook_path, run_summary, publication_tables, raw_tables)

    source_manifest = raw_tables["Raw_Source_Manifest"][0]
    copied_manifest = copy_supporting_sources(final_root, output_root, source_manifest, max_copy_mb=args.max_copy_mb)

    publication_tsv_manifest = pd.DataFrame(publication_tsv_rows)
    publication_tsv_manifest_path = tsv_dir / "publication_ready_tsv_manifest.tsv"
    publication_tsv_manifest.to_csv(publication_tsv_manifest_path, sep="\t", index=False)

    sheet_summary_path = excel_dir / "excel_workbook_sheet_summary.tsv"
    sheet_summary.to_csv(sheet_summary_path, sep="\t", index=False)

    source_manifest_path = source_dir / "supporting_source_file_manifest.tsv"
    if not source_manifest_path.exists():
        copied_manifest.to_csv(source_manifest_path, sep="\t", index=False)

    zip_path = create_zip(output_root, workbook_path)

    report_lines = []
    report_lines.append("PUBLICATION TABLES AND SUPPORTING SOURCE FILES REPORT")
    report_lines.append("")
    report_lines.append(f"Final integrated package root: {final_root}")
    report_lines.append(f"Publication output root: {output_root}")
    report_lines.append("")
    report_lines.append("Generated outputs")
    report_lines.append(f"Combined Excel workbook: {workbook_path}")
    report_lines.append(f"Publication ready TSV folder: {tsv_dir}")
    report_lines.append(f"Supporting source folder: {source_dir}")
    report_lines.append(f"Supporting ZIP bundle: {zip_path}")
    report_lines.append("")
    report_lines.append("Workbook sheets")
    for _, row in sheet_summary.iterrows():
        report_lines.append(f"{row['Sheet name']}: {row['Type']}, rows={row['Rows']}, columns={row['Columns']}")
    report_lines.append("")
    report_lines.append("Source copy summary")
    if copied_manifest.empty:
        report_lines.append("No source copy manifest rows were generated.")
    else:
        report_lines.append(f"Source files copied: {int(copied_manifest['copied'].sum())}")
        report_lines.append(f"Source files skipped or missing: {int((~copied_manifest['copied'].astype(bool)).sum())}")

    report_path = reports_dir / "publication_tables_and_supporting_files_report.txt"
    write_txt(report_path, "\n".join(report_lines))

    summary = {
        "final_integrated_package_root": str(final_root),
        "publication_output_root": str(output_root),
        "combined_excel_workbook": str(workbook_path),
        "publication_ready_tsv_folder": str(tsv_dir),
        "supporting_source_folder": str(source_dir),
        "supporting_zip_bundle": str(zip_path),
        "report": str(report_path),
        "workbook_sheets": int(len(sheet_summary)),
        "publication_ready_tables": int(len(publication_tables)),
        "raw_source_tables": int(len(raw_tables)),
        "source_files_copied": int(copied_manifest["copied"].sum()) if not copied_manifest.empty else 0,
        "source_files_skipped_or_missing": int((~copied_manifest["copied"].astype(bool)).sum()) if not copied_manifest.empty else 0,
    }

    summary_path = output_root / "publication_tables_and_supporting_files_summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print("")
    print("=" * 100)
    print("PUBLICATION TABLES AND SUPPORTING FILES COMPLETE")
    print("=" * 100)
    print("Final integrated package root:", final_root)
    print("Publication output root:", output_root)
    print("Combined Excel workbook:", workbook_path)
    print("Publication ready TSV folder:", tsv_dir)
    print("Supporting source folder:", source_dir)
    print("Supporting ZIP bundle:", zip_path)
    print("Report:", report_path)
    print("Workbook sheets:", summary["workbook_sheets"])
    print("Publication ready tables:", summary["publication_ready_tables"])
    print("Raw source tables:", summary["raw_source_tables"])
    print("Source files copied:", summary["source_files_copied"])
    print("Source files skipped or missing:", summary["source_files_skipped_or_missing"])
    print("")
    print("Publication workbook sheet summary:")
    print(sheet_summary[["Sheet name", "Type", "Rows", "Columns"]].to_string(index=False))

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
